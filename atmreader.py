from turtle import distance
from openpyxl import load_workbook
from scipy import spatial
from constants import LON_COLUMN,LAT_COLUMN,BANK_COLUMN,NET_COLUMN,ADRESS_COLUMN,VISIBILITY,LOADS,PATH,INITIAL_DISTANCE,AtmClass

#Reduce load function: With the index in the atm list of the selected ATM. We can search automatically for that row in the excel
#file, then we can reduce in 1 the load and change the visibility if needed.
def reduce_load(index):
    file = load_workbook(PATH)
    file_sheet =  file["General"]
    load = file_sheet.cell(row = index, column = LOADS)
    load_number = int(load.value)
    if(load_number>0):
        load_number -= 1 
        load.value = load_number
        
    if(load_number == 0):
        visibility = file_sheet.cell(row = index, column = VISIBILITY)
        visibility.value = "False"
        return "False"
    file.save(PATH)
    return "True"

#AtmArray Function: Creates the ATM list with all the Atm classes for later use. It also creates the K Dimensional tree with the
#latitude and longitude of each ATM. With this tree, we can later search for the nearest location in Binary Tree search complexity time
def atmArray():
    file = load_workbook(PATH)
    file_sheet =  file["General"]
    atm_list = []
    kd_tree = []

    #The first row is header, so we start at 2
    for row in range(2, file_sheet.max_row + 1):
        lon = float((file_sheet.cell(row = row, column = LON_COLUMN)).value)
        lat = float((file_sheet.cell(row = row, column = LAT_COLUMN)).value)
        bank = (file_sheet.cell(row = row, column = BANK_COLUMN)).value
        net = (file_sheet.cell(row = row, column = NET_COLUMN)).value
        adress = (file_sheet.cell(row = row, column = ADRESS_COLUMN)).value
        visibility = bool((file_sheet.cell(row = row, column = VISIBILITY)).value)
        index = row
        coordinates = (lon,lat)
        kd_tree.append(coordinates)
        atm = AtmClass(lat,lon,net,bank,adress,visibility,INITIAL_DISTANCE,index)
        atm_list.append(atm)

    tree = spatial.KDTree(kd_tree)
    file.close()
    return tree,atm_list

atmArray()